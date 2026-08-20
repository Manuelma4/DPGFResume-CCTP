"""Banc d'essai CCTP -> DPGF.

Compare ce que le pipeline produit à partir d'un CCTP avec le DPGF réellement
livré par l'économiste pour le même lot. Sans cette mesure, toute affirmation
du type « l'extraction s'est améliorée » reste une opinion.

Organisation attendue du corpus (un sous-dossier par couple) :

    corpus/
      01 ORCHIES/
        CCTP LOT 01 - VRD....docx        <- source
        DPGF Lot 01 - VRD....xlsx        <- référence
      02 NORAUTO LIMOGES/
        ...

Usage :

    .venv\\Scripts\\python.exe tools\\benchmark.py <corpus> [--json rapport.json]
    .venv\\Scripts\\python.exe tools\\benchmark.py <corpus> --detail       # ligne à ligne
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import warnings
from dataclasses import dataclass, field
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
warnings.filterwarnings("ignore", module="openpyxl")

from openpyxl import load_workbook  # noqa: E402

from app.extractors import extract_document  # noqa: E402
from app.parser import _normalized, classify_lot_family, parse_document  # noqa: E402


REFERENCE_CODE = re.compile(r"^\d+(?:\.\d+)*$")
# Une colonne de codes se reconnaît aux codes hiérarchiques ("2.1.1"), jamais
# aux entiers nus : la colonne "Quantité" en est pleine (1, 4, 140...) et
# gagnerait la détection si on comptait tout ce qui ressemble à un nombre.
REFERENCE_DOTTED_CODE = re.compile(r"^\d+(?:\.\d+)+$")
# Lignes de synthèse du classeur : elles portent un montant, jamais un ouvrage.
TOTAL_ROW = re.compile(
    r"^(?:sous[\s-]?total|total|montant|tva|prorata|net\b|s/total)", re.IGNORECASE
)
_REFERENCE_NOISE = {
    "designation",
    "descriptif des ouvrages",
    "description des ouvrages",
    "unite",
    "quantite",
    "code",
}

# Les DPGF réels n'écrivent pas les unités comme le modèle interne : "m2" pour
# m², "m" pour ml, "u" minuscule, et surtout "Ft" (forfait) là où l'application
# produit "Ens". On compare donc des classes d'unité, pas des chaînes — sinon
# on mesurerait une convention typographique au lieu d'une erreur de métré.
UNIT_CLASS = {
    "m²": "surface", "m2": "surface", "m 2": "surface",
    "m³": "volume", "m3": "volume",
    "ml": "longueur", "m": "longueur", "mL": "longueur",
    "u": "unite", "U": "unite", "unité": "unite", "unite": "unite",
    "kg": "masse", "t": "masse",
    "ft": "forfait", "forfait": "forfait", "ens": "forfait",
    "ensemble": "forfait", "forfait/an": "forfait", "f": "forfait",
    "pm": "memoire",
    "h": "heure", "heure": "heure",
    "mois": "duree",
}

# Une désignation générée est considérée comme retrouvée dans la référence
# au-dessus de ce score. Volontairement permissif : l'économiste reformule
# ("Abatage" / "Abattage", "Dépose de clôture existante" / "Dépose des
# clôtures"), et on veut mesurer la couverture du poste, pas la copie exacte.
MATCH_THRESHOLD = 0.62

_STOPWORDS = {
    "de", "du", "des", "le", "la", "les", "un", "une", "et", "en", "a", "au",
    "aux", "pour", "sur", "par", "avec", "d", "l", "the", "ou",
}


def unit_class(value: str) -> str:
    raw = str(value or "").strip()
    return UNIT_CLASS.get(raw, UNIT_CLASS.get(raw.casefold(), "autre"))


def tokens(value: str) -> set[str]:
    words = re.findall(r"[a-z0-9]+", _normalized(value))
    return {word for word in words if word not in _STOPWORDS and len(word) > 1}


def similarity(left: str, right: str) -> float:
    """Mélange similarité de séquence et recouvrement de vocabulaire.

    La seule séquence pénalise trop les inversions ("Dépose de bordure" vs
    "Bordures : dépose"), le seul recouvrement accepte trop de faux positifs
    entre postes voisins d'un même chapitre. La moyenne des deux sépare
    correctement sur le corpus réel."""
    a, b = _normalized(left), _normalized(right)
    if not a or not b:
        return 0.0
    sequence = SequenceMatcher(None, a, b).ratio()
    ta, tb = tokens(left), tokens(right)
    overlap = len(ta & tb) / len(ta | tb) if (ta | tb) else 0.0
    return (sequence + overlap) / 2


@dataclass
class ReferenceLine:
    code: str
    designation: str
    unit: str
    kind: str


@dataclass
class PairResult:
    name: str
    cctp: str
    dpgf: str
    lot_code: str = ""
    lot_title: str = ""
    lot_family: str = ""
    perimeter_method: str = ""
    perimeter_confidence: float = 0.0
    reference_items: int = 0
    generated_items: int = 0
    matched: int = 0
    unit_ok: int = 0
    unit_source_counts: dict[str, int] = field(default_factory=dict)
    flagged: int = 0
    unmatched_reference: list[str] = field(default_factory=list)
    unmatched_generated: list[str] = field(default_factory=list)
    unit_errors: list[dict[str, str]] = field(default_factory=list)
    error: str = ""

    @property
    def recall(self) -> float:
        return self.matched / self.reference_items if self.reference_items else 0.0

    @property
    def precision(self) -> float:
        return self.matched / self.generated_items if self.generated_items else 0.0

    @property
    def unit_accuracy(self) -> float:
        return self.unit_ok / self.matched if self.matched else 0.0

    @property
    def f1(self) -> float:
        total = self.recall + self.precision
        return 2 * self.recall * self.precision / total if total else 0.0


def read_reference(path: Path) -> list[ReferenceLine]:
    """Lit un DPGF livré.

    La colonne des codes est détectée (elle glisse d'un modèle à l'autre : A
    chez ORCHIES/NORAUTO, B chez KEOLIS). Ce qui définit un poste chiffrable
    n'est PAS le code mais la présence d'une unité : les économistes éclatent
    régulièrement un article en sous-lignes numérotées par lettre
    ("2.5.10 Marquage au sol" puis "A - Marquage place PMR | u",
    "B - Marquage place électrique | u"), et ces sous-lignes sont les vraies
    lignes de prix. Les filtrer sur un code hiérarchique en perdait ici une
    cinquantaine et faisait passer pour des inventions les lignes que
    l'application produit correctement."""
    workbook = load_workbook(path, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    best_column, best_score = 1, 0
    for column in range(1, min(worksheet.max_column, 6) + 1):
        score = sum(
            1
            for row in range(1, worksheet.max_row + 1)
            if REFERENCE_DOTTED_CODE.match(
                str(worksheet.cell(row, column).value or "").strip()
            )
        )
        if score > best_score:
            best_column, best_score = column, score

    header_row = next(
        (
            row
            for row in range(1, worksheet.max_row + 1)
            if str(worksheet.cell(row, best_column).value or "").strip().casefold()
            == "code"
        ),
        0,
    )

    lines: list[ReferenceLine] = []
    for row in range(header_row + 1, worksheet.max_row + 1):
        code = str(worksheet.cell(row, best_column).value or "").strip()
        designation = re.sub(
            r"\s+", " ", str(worksheet.cell(row, best_column + 1).value or "").strip()
        )
        unit = str(worksheet.cell(row, best_column + 2).value or "").strip()
        if len(designation) < 3:
            continue
        normalized = _normalized(designation)
        if TOTAL_ROW.match(normalized) or normalized in _REFERENCE_NOISE:
            continue
        lines.append(
            ReferenceLine(
                # Les puces décoratives ("- ", "> ") des sous-lignes ne font
                # pas partie de la désignation et fausseraient l'appariement.
                code=code,
                designation=designation.lstrip("->• ").strip(),
                unit=unit,
                kind="item" if unit else "section",
            )
        )
    return lines


def match_lines(
    generated: list[dict[str, Any]], reference: list[ReferenceLine]
) -> tuple[list[tuple[dict, ReferenceLine, float]], list[dict], list[ReferenceLine]]:
    """Appariement glouton par score décroissant, une référence servant au plus
    une fois — sinon un poste générique ("Divers") capterait plusieurs lignes de
    référence et gonflerait artificiellement le rappel."""
    scored = [
        (similarity(item.get("designation", ""), ref.designation), index, ref_index)
        for index, item in enumerate(generated)
        for ref_index, ref in enumerate(reference)
    ]
    scored.sort(reverse=True)
    used_generated: set[int] = set()
    used_reference: set[int] = set()
    pairs: list[tuple[dict, ReferenceLine, float]] = []
    for score, index, ref_index in scored:
        if score < MATCH_THRESHOLD:
            break
        if index in used_generated or ref_index in used_reference:
            continue
        used_generated.add(index)
        used_reference.add(ref_index)
        pairs.append((generated[index], reference[ref_index], score))
    return (
        pairs,
        [item for index, item in enumerate(generated) if index not in used_generated],
        [ref for index, ref in enumerate(reference) if index not in used_reference],
    )


def evaluate_pair(folder: Path) -> PairResult:
    sources = [
        path
        for path in folder.iterdir()
        if path.suffix.casefold() in {".docx", ".pdf"} and not path.name.startswith("~")
    ]
    references = [
        path
        for path in folder.iterdir()
        if path.suffix.casefold() == ".xlsx" and not path.name.startswith("~")
    ]
    result = PairResult(
        name=folder.name,
        cctp=sources[0].name if sources else "",
        dpgf=references[0].name if references else "",
    )
    if not sources or not references:
        result.error = "couple CCTP/DPGF incomplet"
        return result

    try:
        document = extract_document(sources[0])
        lot = parse_document(document, "bench")
    except Exception as exc:  # noqa: BLE001 - le banc doit survivre à un cas cassé
        result.error = f"{type(exc).__name__}: {exc}"
        return result

    reference = [line for line in read_reference(references[0]) if line.kind == "item"]
    generated = [line for line in lot["lines"] if line.get("kind") == "item"]

    result.lot_code = str(lot.get("code") or "")
    result.lot_title = str(lot.get("title") or "")
    result.lot_family = classify_lot_family(result.lot_title)[0]
    result.perimeter_method = str(lot["perimeter"].get("method") or "")
    result.perimeter_confidence = float(lot["perimeter"].get("confidence") or 0)
    result.reference_items = len(reference)
    result.generated_items = len(generated)
    result.flagged = sum(
        1 for line in generated if line.get("review_status") != "validated"
    )
    counts: dict[str, int] = {}
    for line in generated:
        key = str(line.get("unit_source") or "?")
        counts[key] = counts.get(key, 0) + 1
    result.unit_source_counts = counts

    pairs, unmatched_generated, unmatched_reference = match_lines(generated, reference)
    result.matched = len(pairs)
    for item, ref, _score in pairs:
        if unit_class(str(item.get("unit") or "")) == unit_class(ref.unit):
            result.unit_ok += 1
        else:
            result.unit_errors.append(
                {
                    "designation": ref.designation[:70],
                    "attendu": ref.unit,
                    "obtenu": str(item.get("unit") or ""),
                    "source": str(item.get("unit_source") or ""),
                }
            )
    result.unmatched_reference = [line.designation[:80] for line in unmatched_reference]
    result.unmatched_generated = [
        str(line.get("designation") or "")[:80] for line in unmatched_generated
    ]
    return result


def percent(value: float) -> str:
    return f"{value * 100:>3.0f}%"


def report(results: list[PairResult], detail: bool) -> None:
    print("=" * 100)
    print("BANC D'ESSAI CCTP -> DPGF")
    print("=" * 100)
    header = (
        f"{'dossier':<22} {'lot':<8} {'famille':<14} "
        f"{'réf':>4} {'gén':>4} {'trouvés':>7} {'rappel':>7} {'préc.':>6} "
        f"{'unité':>6} {'signalés':>8}"
    )
    print(header)
    print("-" * 100)
    for result in results:
        if result.error:
            print(f"{result.name:<22} ERREUR: {result.error}")
            continue
        print(
            f"{result.name[:21]:<22} {result.lot_code:<8} {result.lot_family[:13]:<14} "
            f"{result.reference_items:>4} {result.generated_items:>4} "
            f"{result.matched:>7} {percent(result.recall):>7} "
            f"{percent(result.precision):>6} {percent(result.unit_accuracy):>6} "
            f"{result.flagged:>8}"
        )

    valid = [result for result in results if not result.error]
    if valid:
        total_reference = sum(result.reference_items for result in valid)
        total_generated = sum(result.generated_items for result in valid)
        total_matched = sum(result.matched for result in valid)
        total_unit_ok = sum(result.unit_ok for result in valid)
        total_flagged = sum(result.flagged for result in valid)
        recall = total_matched / total_reference if total_reference else 0
        precision = total_matched / total_generated if total_generated else 0
        unit_accuracy = total_unit_ok / total_matched if total_matched else 0
        print("-" * 100)
        print(
            f"{'TOTAL':<22} {'':<8} {'':<14} "
            f"{total_reference:>4} {total_generated:>4} {total_matched:>7} "
            f"{percent(recall):>7} {percent(precision):>6} "
            f"{percent(unit_accuracy):>6} {total_flagged:>8}"
        )
        print()
        print(
            f"  rappel   {percent(recall)}  — postes du DPGF réel effectivement retrouvés"
        )
        print(
            f"  précision{percent(precision)}  — lignes générées qui existent vraiment"
        )
        print(
            f"  unité    {percent(unit_accuracy)}  — unité correcte parmi les postes retrouvés"
        )

    if not detail:
        return

    for result in results:
        if result.error:
            continue
        print()
        print("=" * 100)
        print(f"{result.name}  —  {result.cctp}")
        print(
            f"  lot détecté : {result.lot_code!r} / {result.lot_title!r} "
            f"(famille {result.lot_family})"
        )
        print(
            f"  périmètre   : {result.perimeter_method} "
            f"(confiance {result.perimeter_confidence})"
        )
        print(f"  unit_source : {result.unit_source_counts}")
        if result.unit_errors:
            print(f"  --- unités fausses ({len(result.unit_errors)}) ---")
            for error in result.unit_errors[:15]:
                print(
                    f"    {error['designation']:<72} attendu={error['attendu']:<8} "
                    f"obtenu={error['obtenu']:<5} ({error['source']})"
                )
        if result.unmatched_reference:
            print(
                f"  --- postes du DPGF réel NON générés "
                f"({len(result.unmatched_reference)}) ---"
            )
            for line in result.unmatched_reference[:25]:
                print(f"    - {line}")
        if result.unmatched_generated:
            print(
                f"  --- lignes générées SANS équivalent réel "
                f"({len(result.unmatched_generated)}) ---"
            )
            for line in result.unmatched_generated[:25]:
                print(f"    + {line}")


def main() -> int:
    parser = argparse.ArgumentParser(description="Banc d'essai CCTP -> DPGF")
    parser.add_argument("corpus", type=Path)
    parser.add_argument("--json", type=Path, default=None)
    parser.add_argument("--detail", action="store_true")
    arguments = parser.parse_args()

    if not arguments.corpus.is_dir():
        print(f"Corpus introuvable : {arguments.corpus}", file=sys.stderr)
        return 2

    folders = sorted(
        path
        for path in arguments.corpus.iterdir()
        if path.is_dir() and any(path.iterdir())
    )
    if not folders:
        print(f"Aucun couple dans {arguments.corpus}", file=sys.stderr)
        return 2

    results = [evaluate_pair(folder) for folder in folders]
    report(results, arguments.detail)

    if arguments.json:
        arguments.json.write_text(
            json.dumps(
                [
                    {
                        **{
                            key: value
                            for key, value in vars(result).items()
                            if not key.startswith("_")
                        },
                        "recall": round(result.recall, 4),
                        "precision": round(result.precision, 4),
                        "unit_accuracy": round(result.unit_accuracy, 4),
                        "f1": round(result.f1, 4),
                    }
                    for result in results
                ],
                ensure_ascii=False,
                indent=2,
            ),
            encoding="utf-8",
        )
        print(f"\nRapport JSON : {arguments.json}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
