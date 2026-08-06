from __future__ import annotations

import re
from copy import copy
from datetime import date
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils.units import pixels_to_EMU
from PIL import Image as PillowImage

from . import config


NOTICE = """⚠ NOTICE DE REMPLISSAGE DU DPGF (FICHIER EXCEL)

Avant toute saisie, les entreprises sont invitées à respecter les règles suivantes :

1. Saisie des quantités et des prix
Renseigner ou confirmer uniquement les cellules jaunes prévues pour les quantités et les prix unitaires. Les cellules contenant des formules ne doivent pas être modifiées.

2. Structure du document
Le classeur reste modifiable. Des lignes peuvent être ajoutées si nécessaire ; veillez alors à recopier les formules de quantité, prix et montant ainsi qu'à vérifier les totaux.

3. Options et variantes
Les options ou variantes identifiées doivent être chiffrées sur les lignes dédiées. Elles ne sont pas intégrées dans l'offre de base, sauf indication contraire.

4. Calcul des montants
Les montants sont exprimés en euros hors taxes (€ HT). Les calculs suivent la structure :
Total HT hors prorata → Prorata → Total HT compris prorata → TVA → Total TTC.

Toute modification du fichier pouvant altérer les calculs pourra entraîner l'irrégularité de l'offre."""


def _safe_sheet_title(value: str, used: set[str]) -> str:
    value = re.sub(r"[\[\]:*?/\\]", " ", value)
    value = re.sub(r"\s+", " ", value).strip() or "LOT"
    value = value[:31]
    base = value
    counter = 2
    while value.casefold() in used:
        suffix = f" ({counter})"
        value = f"{base[:31 - len(suffix)]}{suffix}"
        counter += 1
    used.add(value.casefold())
    return value


def _copy_row_style(source, target, source_row: int, target_row: int) -> None:
    for column in range(2, 8):
        source_cell = source.cell(source_row, column)
        target_cell = target.cell(target_row, column)
        target_cell._style = copy(source_cell._style)
        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.fill = copy(source_cell.fill)
            target_cell.border = copy(source_cell.border)
            target_cell.alignment = copy(source_cell.alignment)
            target_cell.protection = copy(source_cell.protection)
            target_cell.number_format = source_cell.number_format


def _sum_formula(rows: list[int]) -> str:
    if not rows:
        return "=0"
    if len(rows) == 1:
        return f"=G{rows[0]}"
    consecutive = all(rows[index] + 1 == rows[index + 1] for index in range(len(rows) - 1))
    if consecutive:
        return f"=SUM(G{rows[0]}:G{rows[-1]})"
    return "=SUM(" + ",".join(f"G{row}" for row in rows) + ")"


def _line_depth(line: dict[str, Any], fallback: int) -> int:
    code = str(line.get("code") or "").strip().strip(".")
    if re.fullmatch(r"[A-Z]?\d+(?:[.\-]\d+)*", code, re.IGNORECASE):
        return len(re.split(r"[.\-]", code))
    return max(1, fallback)


def _set_font(
    cell,
    *,
    size: float | None = None,
    bold: bool | None = None,
    italic: bool | None = None,
) -> None:
    font = copy(cell.font)
    if size is not None:
        font.sz = size
    if bold is not None:
        font.b = bold
    if italic is not None:
        font.i = italic
    cell.font = font


def _add_logo(worksheet) -> None:
    logo_path = config.ASSETS_DIR / "moduo_logo.png"
    if not logo_path.exists():
        return
    # Normalise the actual image bytes instead of trusting the extension.
    # A WebP file renamed to .png is otherwise detected by openpyxl as WebP;
    # openpyxl then writes an unsupported .webp media entry and fails while
    # packaging the workbook.
    logo_bytes = BytesIO()
    with PillowImage.open(logo_path) as source:
        source.save(logo_bytes, format="PNG")
    logo_bytes.seek(0)

    # The existing title row remains 51 pt high. The vertical logo fits
    # inside it with a small margin and does not move the centered title.
    logo = ExcelImage(logo_bytes)
    logo.height = 54
    logo.width = 100
    logo.anchor = OneCellAnchor(
        _from=AnchorMarker(
            col=1,
            row=1,
            colOff=pixels_to_EMU(6),
            rowOff=pixels_to_EMU(7),
        ),
        ext=XDRPositiveSize2D(
            cx=pixels_to_EMU(logo.width),
            cy=pixels_to_EMU(logo.height),
        ),
    )
    worksheet.add_image(logo)


def _populate_sheet(
    worksheet,
    style_source,
    lot: dict[str, Any],
    project: dict[str, Any],
    *,
    prorata_rate: float,
    vat_rate: float,
) -> None:
    style_rows = {
        "section": 12,
        "item": 14,
        "subtotal": 21,
        "total_ht": 48,
        "prorata": 49,
        "total_prorata": 50,
        "vat": 51,
        "total_ttc": 52,
        "notice": 55,
    }
    for merged in list(worksheet.merged_cells.ranges):
        if merged.min_row >= 12:
            worksheet.unmerge_cells(str(merged))
    if worksheet.max_row >= 12:
        worksheet.delete_rows(12, worksheet.max_row - 11)
    for row_number in list(worksheet.row_dimensions):
        if row_number >= 12:
            del worksheet.row_dimensions[row_number]

    worksheet["B2"] = "\n".join(
        value for value in (
            str(project.get("name") or "").strip(),
            str(project.get("client") or "").strip(),
            str(project.get("reference") or "").strip(),
        )
        if value
    )
    worksheet["B4"] = str(project.get("phase") or "DCE").upper()
    worksheet["C4"] = "DPGF"
    worksheet["G4"] = date.today()
    worksheet["G4"].number_format = "dd/mm/yyyy"
    worksheet["D6"] = "À compléter"
    _add_logo(worksheet)
    lot_label = "LOT"
    if lot.get("code"):
        lot_label += f" {lot['code']}"
    if lot.get("title"):
        lot_label += f" — {lot['title']}"
    worksheet["B8"] = lot_label
    worksheet["B11"] = "DESCRIPTIF DES OUVRAGES"

    row = 12
    all_included_rows: list[int] = []
    lines = list(lot.get("lines") or [])

    # Build an actual tree from the numeral levels. Only sections may own
    # children; a leaf item at the same level closes the previous branch.
    roots: list[dict[str, Any]] = []
    section_stack: list[dict[str, Any]] = []
    for line in lines:
        level = max(1, int(line.get("level") or 1))
        node = {"line": line, "level": level, "children": []}
        while section_stack and int(section_stack[-1]["level"]) >= level:
            section_stack.pop()
        if section_stack:
            section_stack[-1]["children"].append(node)
        else:
            roots.append(node)
        if str(line.get("kind") or "item") == "section":
            section_stack.append(node)

    def render_item(line: dict[str, Any], level: int) -> list[int]:
        nonlocal row
        depth = _line_depth(line, level)
        title_style = depth <= 2
        _copy_row_style(style_source, worksheet, style_rows["item"], row)
        if title_style:
            # A leaf can still be a principal article (for example 3.6 next
            # to section 3.5). Give its code and designation the same visual
            # hierarchy as a title without merging away unit/price columns.
            for column in (2, 3):
                source_cell = style_source.cell(style_rows["section"], column)
                target_cell = worksheet.cell(row, column)
                target_cell._style = copy(source_cell._style)
                target_cell.font = copy(source_cell.font)
                target_cell.fill = copy(source_cell.fill)
                target_cell.border = copy(source_cell.border)
                target_cell.protection = copy(source_cell.protection)
        worksheet.cell(row, 2, str(line.get("code") or ""))
        designation = str(line.get("designation") or "")
        if not bool(line.get("included", True)):
            designation = f"{designation} (OPTION / VARIANTE)"
        worksheet.cell(row, 3, designation.upper() if title_style else designation)
        worksheet.cell(row, 4, str(line.get("unit") or "Ens"))
        worksheet.cell(row, 5, line.get("quantity"))
        worksheet.cell(row, 6, line.get("unit_price"))
        worksheet.cell(row, 7, f'=IF(OR(E{row}="",F{row}=""),"",E{row}*F{row})')
        for column in range(2, 8):
            _set_font(worksheet.cell(row, column), size=8, italic=False)
        if title_style:
            for column in (2, 3):
                _set_font(
                    worksheet.cell(row, column),
                    size=9,
                    bold=True,
                    italic=False,
                )
        else:
            _set_font(worksheet.cell(row, 3), size=8, bold=False, italic=True)
        source_page = line.get("source_page")
        source_excerpt = str(line.get("source_excerpt") or "").strip()
        if source_page or source_excerpt:
            source_label = (
                f"Source : page {source_page}"
                if source_page
                else "Source : document Word"
            )
            if source_excerpt:
                source_label += f"\n\n{source_excerpt[:900]}"
            worksheet.cell(row, 3).comment = Comment(
                source_label, "DPGFResume CCTP"
            )
        unit_source = str(line.get("unit_source") or "")
        if unit_source == "explicit":
            unit_note = "Unité explicitement mentionnée dans le CCTP."
        elif unit_source == "rule":
            unit_note = "Unité proposée selon la prestation ; à confirmer."
        elif unit_source == "default":
            unit_note = "Unité forfaitaire proposée par défaut ; à confirmer."
        else:
            unit_note = "Unité renseignée manuellement."
        worksheet.cell(row, 4).comment = Comment(unit_note, "DPGFResume CCTP")
        if line.get("quantity") is None:
            worksheet.cell(row, 5).comment = Comment(
                "Quantité non démontrée par le CCTP : cellule volontairement vide.",
                "DPGFResume CCTP",
            )
        for input_column in (5, 6):
            worksheet.cell(row, input_column).fill = PatternFill(
                "solid", fgColor="FFF2CC"
            )
            worksheet.cell(row, input_column).protection = Protection(locked=False)
        worksheet.cell(row, 5).number_format = "#,##0.00"
        worksheet.cell(row, 6).number_format = '#,##0.00\\ "€"'
        worksheet.cell(row, 7).number_format = '#,##0.00\\ "€"'
        worksheet.cell(row, 3).alignment = Alignment(
            horizontal="left",
            vertical="center" if title_style else "top",
            wrap_text=True,
            indent=max(0, min(8, level - 2)),
        )
        worksheet.row_dimensions[row].height = max(
            19 if title_style else 18,
            15 * min(4, designation.count("\n") + 1),
        )
        item_row = row
        included_rows = [item_row] if bool(line.get("included", True)) else []
        if included_rows:
            all_included_rows.extend(included_rows)
        row += 1
        if depth == 2:
            # An x.x line can be both the title and the priced object. Keep
            # its unit/price cells and close it with its own subtotal.
            _copy_row_style(style_source, worksheet, style_rows["subtotal"], row)
            code = str(line.get("code") or "")
            worksheet.cell(row, 3, f"Sous-total {code}".strip())
            worksheet.cell(row, 3).alignment = Alignment(
                horizontal="right",
                vertical="center",
            )
            worksheet.cell(row, 7, _sum_formula(included_rows))
            worksheet.row_dimensions[row].height = 18
            row += 1
        return included_rows

    def is_descriptor_section(line: dict[str, Any]) -> bool:
        designation = str(line.get("designation") or "").casefold()
        return any(
            marker in designation
            for marker in (
                "description des ouvrages",
                "description des travaux",
                "descriptif des ouvrages",
            )
        )

    def render_node(node: dict[str, Any]) -> list[int]:
        nonlocal row
        line = node["line"]
        level = int(node["level"])
        depth = _line_depth(line, level)
        if str(line.get("kind") or "item") != "section":
            return render_item(line, level)

        if depth <= 2:
            _copy_row_style(style_source, worksheet, style_rows["section"], row)
            worksheet.cell(row, 2, str(line.get("code") or ""))
            worksheet.cell(row, 3, str(line.get("designation") or "").upper())
            worksheet.merge_cells(
                start_row=row, start_column=3, end_row=row, end_column=7
            )
            for column in range(2, 8):
                _set_font(
                    worksheet.cell(row, column),
                    size=9,
                    bold=True,
                    italic=False,
                )
            worksheet.cell(row, 3).alignment = Alignment(
                horizontal="left",
                vertical="center",
                wrap_text=True,
                indent=max(0, min(8, depth - 1)),
            )
        else:
            # Nested numerals (x.x.x and deeper) are contextual objects, not
            # autonomous titles. They stay visible but never create a subtotal.
            _copy_row_style(style_source, worksheet, style_rows["item"], row)
            worksheet.cell(row, 2, str(line.get("code") or ""))
            worksheet.cell(row, 3, str(line.get("designation") or ""))
            for column in range(2, 8):
                _set_font(
                    worksheet.cell(row, column),
                    size=8,
                    bold=False,
                    italic=False,
                )
            _set_font(worksheet.cell(row, 3), size=8, bold=False, italic=True)
            worksheet.cell(row, 3).alignment = Alignment(
                horizontal="left",
                vertical="top",
                wrap_text=True,
                indent=max(0, min(8, depth - 2)),
            )
        worksheet.row_dimensions[row].height = 19
        row += 1

        branch_item_rows: list[int] = []
        for child in node["children"]:
            branch_item_rows.extend(render_node(child))

        if depth == 2 and branch_item_rows and not is_descriptor_section(line):
            _copy_row_style(style_source, worksheet, style_rows["subtotal"], row)
            code = str(line.get("code") or "")
            worksheet.cell(row, 3, f"Sous-total {code}".strip())
            worksheet.cell(row, 3).alignment = Alignment(
                horizontal="right",
                vertical="center",
                indent=max(0, min(8, level - 2)),
            )
            worksheet.cell(row, 7, _sum_formula(branch_item_rows))
            worksheet.row_dimensions[row].height = 18
            row += 1
        return branch_item_rows

    for root in roots:
        render_node(root)

    if not lines:
        _copy_row_style(style_source, worksheet, style_rows["item"], row)
        worksheet.cell(row, 3, "Aucun poste extrait — à compléter")
        row += 2

    row += 1
    total_ht_row = row
    for key, label in (
        ("total_ht", "TOTAL € H.T. hors prorata"),
        ("prorata", f"Prorata {prorata_rate:g}%"),
        ("total_prorata", "TOTAL € H.T. compris prorata"),
        ("vat", f"TVA à {vat_rate:.2f} %"),
        ("total_ttc", "TOTAL € T.T.C."),
    ):
        _copy_row_style(style_source, worksheet, style_rows[key], row)
        worksheet.cell(row, 3, label)
        worksheet.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
        row += 1

    # The total always references every included item exactly once. Displayed
    # nested subtotals are informative and can never double-count the total.
    worksheet.cell(total_ht_row, 7, _sum_formula(all_included_rows))
    worksheet.cell(total_ht_row + 1, 7, f"=G{total_ht_row}*{prorata_rate}/100")
    worksheet.cell(total_ht_row + 2, 7, f"=SUM(G{total_ht_row}:G{total_ht_row + 1})")
    worksheet.cell(total_ht_row + 3, 7, f"=G{total_ht_row + 2}*{vat_rate}/100")
    worksheet.cell(total_ht_row + 4, 7, f"=SUM(G{total_ht_row + 2},G{total_ht_row + 3})")
    for formula_row in range(total_ht_row, total_ht_row + 5):
        worksheet.cell(formula_row, 7).number_format = '#,##0.00\\ "€"'

    row += 2
    _copy_row_style(style_source, worksheet, style_rows["notice"], row)
    worksheet.cell(row, 2, NOTICE)
    worksheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
    worksheet.cell(row, 2).alignment = Alignment(
        horizontal="left", vertical="top", wrap_text=True
    )
    worksheet.cell(row, 2).font = Font(
        name="Aptos", size=9, color="44546A", bold=False
    )
    worksheet.cell(row, 2).fill = PatternFill("solid", fgColor="F2F6FA")
    worksheet.row_dimensions[row].height = 245

    worksheet.freeze_panes = "B10"
    # The source template keeps its first sheet scrolled to row 9. Reset the
    # saved viewport so every generated lot opens with the project header
    # (rows 1–8) visible, exactly like the copied sheets.
    worksheet.sheet_view.topLeftCell = "A1"
    for header_row in range(1, 10):
        worksheet.row_dimensions[header_row].hidden = False
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.page_setup.orientation = "portrait"
    worksheet.print_area = f"B1:G{row}"
    worksheet.print_title_rows = "1:9"
    # Keep the generated workbook fully editable so users can add, remove,
    # move or format lines directly in Excel.
    worksheet.protection.sheet = False
    worksheet["D6"].protection = Protection(locked=False)
    worksheet.sheet_view.showGridLines = False

    for column, width in {"B": 12, "C": 58, "D": 10, "E": 12, "F": 14, "G": 17}.items():
        worksheet.column_dimensions[column].width = width
    worksheet.auto_filter.ref = None


def export_analysis(
    analysis: dict[str, Any],
    destination: Path,
    *,
    prorata_rate: float = 1.5,
    vat_rate: float = 20.0,
) -> Path:
    if not config.TEMPLATE_PATH.exists():
        raise FileNotFoundError("Le modèle DPGF TYPE.xlsx est absent")
    lots = list(analysis.get("lots") or [])
    if not lots:
        raise ValueError("Aucun lot ne peut être exporté")
    workbook = load_workbook(config.TEMPLATE_PATH)
    try:
        workbook.defined_names.clear()
    except AttributeError:
        pass
    base = workbook.active
    sheets = [base]
    for _ in lots[1:]:
        sheets.append(workbook.copy_worksheet(base))

    used_titles: set[str] = set()
    for worksheet, lot in zip(sheets, lots):
        style_source = workbook.copy_worksheet(worksheet)
        try:
            label = f"LOT {lot.get('code') or ''} {lot.get('title') or ''}".strip()
            worksheet.title = _safe_sheet_title(label, used_titles)
            _populate_sheet(
                worksheet,
                style_source,
                lot,
                analysis.get("project") or {},
                prorata_rate=prorata_rate,
                vat_rate=vat_rate,
            )
        finally:
            workbook.remove(style_source)

    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)
    return destination


def safe_export_name(project: dict[str, Any]) -> str:
    base = str(project.get("reference") or project.get("name") or "Projet")
    base = re.sub(r"[^\w\-]+", "_", base, flags=re.UNICODE).strip("_")
    return f"DPGF_{base or 'Projet'}.xlsx"


def safe_lot_export_name(
    project: dict[str, Any],
    lot: dict[str, Any],
    used: set[str] | None = None,
) -> str:
    project_name = Path(safe_export_name(project)).stem
    lot_label = "_".join(
        value
        for value in (
            "LOT",
            str(lot.get("code") or "").strip(),
            str(lot.get("title") or "").strip(),
        )
        if value
    )
    lot_label = re.sub(r"[^\w\-]+", "_", lot_label, flags=re.UNICODE).strip("_")
    base = f"{project_name}_{lot_label or 'LOT'}"[:170].rstrip("_")
    name = f"{base}.xlsx"
    if used is None:
        return name
    counter = 2
    while name.casefold() in used:
        suffix = f"_{counter}"
        name = f"{base[:170 - len(suffix)]}{suffix}.xlsx"
        counter += 1
    used.add(name.casefold())
    return name


def export_analysis_files(
    analysis: dict[str, Any],
    destination_directory: Path,
    *,
    prorata_rate: float = 1.5,
    vat_rate: float = 20.0,
) -> list[Path]:
    lots = list(analysis.get("lots") or [])
    if not lots:
        raise ValueError("Aucun lot ne peut être exporté")
    destination_directory.mkdir(parents=True, exist_ok=True)
    used_names: set[str] = set()
    outputs: list[Path] = []
    for lot in lots:
        destination = destination_directory / safe_lot_export_name(
            analysis.get("project") or {},
            lot,
            used_names,
        )
        export_analysis(
            {**analysis, "lots": [lot]},
            destination,
            prorata_rate=prorata_rate,
            vat_rate=vat_rate,
        )
        outputs.append(destination)
    return outputs


def safe_export_archive_name(project: dict[str, Any]) -> str:
    return f"{Path(safe_export_name(project)).stem}_Lots.zip"
